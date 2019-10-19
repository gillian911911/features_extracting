from scipy.signal.windows import blackmanharris
import pyAudioAnalysis
from pyAudioAnalysis import audioBasicIO as aIO
from pyAudioAnalysis import audioSegmentation as aS
from pyAudioAnalysis import convertToWav
from pyAudioAnalysis import audioFeatureExtraction as aFE
from pyAudioAnalysis import audioVisualization as aV
from pyAudioAnalysis.utilities import load_file

import numpy as np
import pydub
from pydub import AudioSegment
import  os
import matplotlib.pyplot as plt
from IPython.display import Audio
import openpyxl as xl

def Trans_Other_to_Wav (filepath):
    pydub.AudioSegment.converter = "C:/CodeTools/ffmpeg/bin/ffmpeg.exe"
    voice = pydub.AudioSegment.from_mp3(filepath)
    voice.export("new.wav", format="wav")
    return voice




def xlwrite_short(F,name):
    wb = xl.Workbook()
    sheet = wb.active

    column = 1
    for i in F[1]:
        sheet.cell(row=1, column=column, value=i)
        column += 1

    column = 1
    for i in F[0]:
        row = 2
        for k in i:
            # print('第%d次打印：' % (count-1), k)
            sheet.cell(row=row, column=column, value=k)
            row += 1
        column += 1
    wb.save("%s.xlsx" % name)

def xlwrite_mid(F,name):
    wb = xl.Workbook()
    sheet = wb.active
    column = 1
    for i in MF[2]:
        sheet.cell(row = 1, column = column, value  = i)
        column +=1
    column = 1
    for i in MF[0]:#i 1:34
        row = 2
        for k in i: # k range
            sheet.cell(row = row, column =column,value = k)
            row +=1
        column +=1 #column :1-34
    wb.save("%s.xlsx" % name)

# voice = Trans_Other_to_Wav("C:/CodeTools/voice.mp3")
if __name__ == '__main__':
    file_path = "/Users/gillian/PycharmProjects/voice_mining/voice_ming20191018/V_features_extracting/CASIA/CASIA database/liuchanhg/angry"
    file_list = load_file(file_path)
    features_st = list()
    features_mt = list()
    print(file_list[1])
    for i, file in enumerate(file_list):
        if file[-3:] != "wav":
            continue
        [Fs, signal]=aIO.readAudioFile(file)
        #Audio(data = signal, rate = Fs)

        print(type(Fs), type(signal))
        print(Fs)
        print(signal.shape)

        #signal: Mono + normalization
        signal = aIO.stereo2mono(signal)
        signal = np.double(signal)


        Fs1 = 0.05*Fs
        Fs2 = 0.025*Fs
        print("FS1",Fs1)
        print("Fs2",Fs2)
        SF = aFE.stFeatureExtraction(signal, Fs, Fs1, Fs2)#fs:采样频率，win:计算窗口； step:步长
        MF = aFE.mtFeatureExtraction(signal, Fs, 1.0*Fs, 1.0*Fs, 0.05*Fs, 0.025*Fs)

        print("SF1%s" %i,SF[1])
        print("MF00%s" %i, MF[0][0])
        print("MF1%s" %i,MF[1])
        print("MF2%s" %i, MF[2])
        print(len(MF[0]))
        print(len(MF[1]))
        print(len(MF[2]))
        xlwrite_short(SF, "Short_Features_%s" %i)
        xlwrite_mid(MF,"Middle_Features_%s" %i)


        plt.subplot(2, 1, 1)
        plt.plot(SF[0][0])
        plt.xlabel("Frame nomber")
        plt.ylabel("ZCR")

        print(SF[0][2])
        plt.subplot(2,1,2)
        plt.plot(SF[0][1])
        plt.xlabel("Frame nomber")
        plt.ylabel("Energy")
        plt.show()

        # Short_Features_Short = aFE.stFeatureSpeed(signal, Fs, Fs1, Fs2)
          # print(Short_Features_Short) #stHarmonic(x, fs)):return  (HR, f0) f0:Get fundamental frequency:
        print(type(SF[0]))
        Beats= aFE.beatExtraction(SF[0], 0.25, PLOT=False) # -窗长为0.25s
        print("Beats %s" %i ,Beats)
# \
